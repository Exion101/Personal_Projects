from typing import Counter
import openpyxl

# create SI and CI functions

def SimpleInterest( time, interest):
    SI = 1 + time*interest
    return SI
def CompoundInterest( time, interest):
    CI = (1 + interest) ** time
    return CI

wb = openpyxl.Workbook()
sheet = wb.active

# Create headers
func = sheet.cell(row = 1, column = 1)
func.value = "Time"

si_header = sheet.cell(row = 1, column = 2)
si_header.value = "Simple Interest"

ci_header = sheet.cell(row = 1, column = 3)
ci_header.value = "Compound Interest"

# set time intervals
t= [0,0.2,0.4,0.6,0.8,1,5,10,20, 35, 50, 65, 80]
# set interest
i = .036

# blanks start at 2
ct = 2
for time in t:
    #time column
    temp1 = sheet.cell(row = ct, column = 1)
    temp1.value = time
    #si
    temp2 = sheet.cell(row = ct, column= 2)
    temp2.value = SimpleInterest(time, i)
    #ci
    temp3 = sheet.cell(row = ct, column=3)
    temp3.value = CompoundInterest(time,i)

    ct += 1

wb.save("proj1.xlsx")
